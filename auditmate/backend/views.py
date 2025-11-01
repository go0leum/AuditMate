import os
import json
import datetime
import zipfile
import pandas as pd
import numpy as np
import openpyxl
from django.http import JsonResponse, FileResponse, Http404
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from django.views.decorators.http import require_http_methods
import io

from .constants import (
    WORKSHEET_NAME, ALLOWED_EXTENSIONS, METADATA_FILENAME,
    REVIEW_COLUMNS, CORE_COLUMNS, PROGRESS_COLUMN,
    COMMON_HEADER_COLUMNS, Messages, DEFAULT_METADATA_STRUCTURE
)

UPLOAD_DIR = os.path.join(settings.BASE_DIR, "Upload_file")
RULE_DIR = os.path.join(settings.BASE_DIR, "Rule_file")
CONTACT_DIR = os.path.join(settings.BASE_DIR, "Contact_file")

def calculate_progress(worksheet, column_line):
    """
    openpyxl 워크시트에서 '검토사항' 열을 기준으로 진행도를 계산합니다.
    진행도 = 마지막으로 채워진 행의 인덱스 / 전체 행 수 * 100
    """
    # 헤더 행 결정 (columnLine 기준)
    header_row = column_line + 1 if column_line >= 0 else 1
    
    # '검토사항' 컬럼 찾기
    review_col_idx = None
    for col_idx in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if cell_value and str(cell_value).strip() == PROGRESS_COLUMN:
            review_col_idx = col_idx
            break
    
    if review_col_idx is None:
        return 0  # '검토사항' 컬럼이 없으면 0% 반환
    
    # 데이터 행 범위 계산 (헤더 다음 행부터)
    data_start_row = header_row + 1
    total_rows = 0
    last_filled_row = 0
    
    # 실제 데이터가 있는 행들 확인
    for row_idx in range(data_start_row, worksheet.max_row + 1):
        # 해당 행에 데이터가 있는지 확인 (모든 컬럼 체크)
        has_data = False
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and str(cell_value).strip():
                has_data = True
                break
        
        if has_data:
            total_rows += 1
            # '검토사항' 컬럼에 값이 있는지 확인
            review_value = worksheet.cell(row=row_idx, column=review_col_idx).value
            if review_value is not None and str(review_value).strip():
                last_filled_row = total_rows
    
    if total_rows == 0:
        return 0
    
    return round(last_filled_row * 100 / total_rows, 2)

def find_column_line(worksheet):
    """
    openpyxl 워크시트에서 공통 컬럼 중 하나라도 포함된 가장 첫 번째 행(0-based index)을 찾습니다.
    """
    # 처음 10행만 검사
    for row_idx in range(1, min(worksheet.max_row + 1, 11)):
        row_values = []
        
        # 해당 행의 모든 셀 값을 가져와서 문자열로 변환
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_values.append(str(cell_value).strip())
        
        # 정확한 매칭 검사
        matches = [col for col in COMMON_HEADER_COLUMNS if col in row_values]
        if len(matches) >= 3:  # 최소 3개 이상의 컬럼이 매칭되면 헤더로 판단
            return row_idx - 1  # 0-based index로 반환
    
    return -1

def list_files(request):
    if not os.path.exists(UPLOAD_DIR):
        return JsonResponse({"error": Messages.UPLOAD_DIR_NOT_EXISTS}, status=404)

    result = []
    for folder_name in os.listdir(UPLOAD_DIR):
        folder_path = os.path.join(UPLOAD_DIR, folder_name)
        if os.path.isdir(folder_path):  
            xlsx_files = [f for f in os.listdir(folder_path) if f.endswith(ALLOWED_EXTENSIONS[0])]
            xlsx_file = xlsx_files[0] if xlsx_files else None

            metadata_path = os.path.join(folder_path, METADATA_FILENAME)
            metadata_info = {}

            # metadata.json 읽기
            if os.path.exists(metadata_path):
                try:
                    with open(metadata_path, "r", encoding="utf-8") as f:
                        metadata_info = json.loads(f.read())  
                except Exception as e:
                    print(f"Metadata 파일 읽기 오류: {e}")

            # 진행도 계산 
            progress = 0
            if xlsx_file:
                try:
                    # metadata에서 columnLine 정보 가져오기
                    column_line = metadata_info.get("columnLine", 0)
                    
                    # openpyxl로 워크북 로드하여 진행도 계산
                    workbook = openpyxl.load_workbook(os.path.join(folder_path, xlsx_file), data_only=True)
                    
                    # "집행내역" 시트 확인
                    if WORKSHEET_NAME in workbook.sheetnames:
                        worksheet = workbook[WORKSHEET_NAME]
                        progress = calculate_progress(worksheet, column_line)
                    else:
                        # "집행내역" 시트가 없으면 첫 번째 시트 사용
                        if workbook.sheetnames:
                            worksheet = workbook[workbook.sheetnames[0]]
                            progress = calculate_progress(worksheet, column_line)

                    # metadata.json 업데이트
                    metadata_info["progress"] = progress
                    with open(metadata_path, "w", encoding="utf-8") as f:
                        json.dump(metadata_info, f, ensure_ascii=False, indent=4)

                except Exception as e:
                    print(f"Excel 진행도 계산 오류: {e}")

            result.append({
                "folderName": folder_name,
                "xlsxFile": xlsx_file,
                "lastModified": metadata_info.get("lastModified", ""),
                "progress": progress,  # 계산된 진행도 반영
                "ruleName": metadata_info.get("ruleName", ""),
            })

    return JsonResponse(result, safe=False)

def list_rules(request):
    if not os.path.exists(RULE_DIR):
        return JsonResponse({"error": Messages.RULE_DIR_NOT_EXISTS}, status=404)

    result = []
    
    for folder_name in os.listdir(RULE_DIR):
        folder_path = os.path.join(RULE_DIR, folder_name)
        if not os.path.isdir(folder_path):
            continue
            
        rule_info = load_rule_from_folder(folder_path, folder_name)
        result.append(rule_info)

    return JsonResponse(result, safe=False, json_dumps_params={'ensure_ascii': False})

def load_rule_from_folder(folder_path, folder_name):
    """
    개별 규칙 폴더에서 규칙 정보를 로딩하는 함수
    """
    metadata_path = os.path.join(folder_path, METADATA_FILENAME)
    default_upload_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 기본값 설정
    result = {
        "folderName": folder_name,
        "documentRule": {},
        "categoryRule": {},
        "uploadTime": default_upload_time,
    }
    
    if os.path.exists(metadata_path):
        try:
            with open(metadata_path, "r", encoding="utf-8") as f:
                metadata = json.load(f)
            
            # 메타데이터에서 정보 추출
            result["uploadTime"] = metadata.get("uploadTime", default_upload_time)
            result["categoryRule"] = metadata.get("categoryRule", {})
            
            # documentRule 파일 로딩
            document_rule_filename = metadata.get("documentRule")
            if document_rule_filename:
                document_rule_path = os.path.join(folder_path, document_rule_filename)
                if os.path.exists(document_rule_path):
                    with open(document_rule_path, "r", encoding="utf-8") as f:
                        result["documentRule"] = json.load(f)
                else:
                    print(f"Warning: documentRule 파일이 존재하지 않음 - {document_rule_path}")
            
        except Exception as e:
            print(f"Error: 메타데이터 로딩 실패 ({folder_name}): {e}")

    else:
        print(f"Warning: metadata.json이 없음 - {folder_name}")
    
    return result

def download_file(request, folder_name, file_name):
    file_path = os.path.join(UPLOAD_DIR, folder_name, file_name)
    metadata_path = os.path.join(UPLOAD_DIR, folder_name, METADATA_FILENAME)

    if not os.path.exists(file_path):
        raise Http404("File not found")
    
    # 메타데이터에서 columnLine 정보 가져오기
    column_line = 0
    if os.path.exists(metadata_path):
        with open(metadata_path, "r", encoding="utf-8") as f:
            metadata = json.load(f)
        column_line = metadata.get("columnLine", 0)
    else:
        metadata = {}

    # type 파라미터 확인
    download_type = request.GET.get('type', 'full')

    # xlsx 파일이면 type에 따라 열 제거
    if file_name.lower().endswith('.xlsx'):
        try:
            import tempfile
            
            # openpyxl로 워크북 로드
            workbook = openpyxl.load_workbook(file_path)
            
            # "집행내역" 시트 확인
            if WORKSHEET_NAME not in workbook.sheetnames:
                return JsonResponse({"error": Messages.WORKSHEET_NOT_FOUND.format(WORKSHEET_NAME)}, status=400)
            
            worksheet = workbook[WORKSHEET_NAME]
            
            if download_type == 'no_review':
                # '검토사항', '메모' 열 찾아서 제거
                cols_to_exclude = REVIEW_COLUMNS
                
                # 헤더 행 찾기 (column_line 기준)
                header_row = column_line + 1 if column_line > 0 else 1
                
                # 헤더에서 제외할 컬럼의 인덱스 찾기
                cols_to_delete = []
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=header_row, column=col_idx).value
                    if cell_value and str(cell_value).strip() in cols_to_exclude:
                        cols_to_delete.append(col_idx)
                
                # 뒤에서부터 컬럼 삭제 (인덱스가 변경되지 않도록)
                for col_idx in reversed(cols_to_delete):
                    worksheet.delete_cols(col_idx)
            
            # 임시 파일로 저장
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                workbook.save(tmp.name)
                response = FileResponse(open(tmp.name, "rb"), as_attachment=True, filename=file_name)
            return response
        except Exception as e:
            return JsonResponse({"error": Messages.EXCEL_EXPORT_ERROR.format(str(e))}, status=500)

    if os.path.isdir(file_path):
        # 디렉토리라면 zip으로 압축해서 반환
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for root, _, files in os.walk(file_path):
                for f in files:
                    abs_path = os.path.join(root, f)
                    rel_path = os.path.relpath(abs_path, file_path)
                    zip_file.write(abs_path, arcname=rel_path)
        zip_buffer.seek(0)
        zip_filename = f"{file_name}.zip"
        return FileResponse(zip_buffer, as_attachment=True, filename=zip_filename)
    else:
        return FileResponse(open(file_path, "rb"), as_attachment=True, filename=file_name)

@csrf_exempt
def upload_files(request):
    if request.method == "POST":
        upload_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        upload_dir = os.path.join(UPLOAD_DIR, upload_time)

        metadata = {
            "folderName": upload_time,
            "xlsxFile": None,
            "lastModified": upload_time,
            "progress": 0,
            "ruleName": None,
            "columnLine": 0,
        }
        
        os.makedirs(upload_dir, exist_ok=True)

        fs = FileSystemStorage(location=upload_dir)

        # Excel 파일 저장
        if "excel_file" in request.FILES:
            excel_file = request.FILES["excel_file"]
            saved_excel_path = fs.save(excel_file.name, excel_file)
            metadata["xlsxFile"] = saved_excel_path

            # 진행도 계산 (Excel 파일에서 "검토사항" 열 기준), 컬럼 행 찾기
            try:
                # openpyxl로 워크북 로드하여 컬럼 라인 찾기
                workbook = openpyxl.load_workbook(os.path.join(upload_dir, metadata["xlsxFile"]), data_only=True)
                
                # "집행내역" 시트 확인
                if WORKSHEET_NAME not in workbook.sheetnames:
                    return JsonResponse({
                        "status": "error",
                        "message": Messages.WORKSHEET_NOT_FOUND.format(WORKSHEET_NAME),
                        "error_detail": "집행내역 시트 없음"
                    }, status=400)
                
                worksheet = workbook[WORKSHEET_NAME]
                metadata["columnLine"] = find_column_line(worksheet)
                
                # openpyxl로 진행도 계산
                metadata["progress"] = calculate_progress(worksheet, metadata["columnLine"])
                metadata["lastModified"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            except Exception as e:
                # 정렬/필터 등으로 읽기 실패 시 에러 메시지 반환
                return JsonResponse({
                    "status": "error",
                    "message": Messages.UPLOAD_FAILED,
                    "error_detail": str(e)
                }, status=400)

        # Rule_file에서 첫번째 폴더 이름 가져오기
        rule_name = None
        if os.path.exists(RULE_DIR):
            rule_folders = [f for f in os.listdir(RULE_DIR) if os.path.isdir(os.path.join(RULE_DIR, f))]
            if rule_folders:
                rule_name = rule_folders[0]
            else:
                rule_name = None  # 폴더가 없으면 None
        else:
            rule_name = None  # Rule_file 폴더가 없으면 None
        
        metadata["ruleName"] = rule_name

        # metadata.json 생성
        with open(os.path.join(upload_dir, METADATA_FILENAME), "w", encoding="utf-8") as f:
            json.dump(metadata, f, ensure_ascii=False, indent=4)

        return JsonResponse({"message": Messages.UPLOAD_SUCCESS, "metadata": metadata})

    return JsonResponse({"error": Messages.INVALID_REQUEST_METHOD}, status=400)


@csrf_exempt
def read_xlsx(request):
    # 기본적으로 항상 있어야 하는 핵심 컬럼들만 정의
    core_columns = CORE_COLUMNS

    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            folderName = body.get('folderName')
            xlsxFile = body.get('xlsxFile')

            if not folderName:
                return JsonResponse({'status': 'error', 'message': Messages.MISSING_FILENAME}, status=400)

            file_path = os.path.join(UPLOAD_DIR, folderName, xlsxFile)
            metadata_path = os.path.join(UPLOAD_DIR, folderName, METADATA_FILENAME)

            if not os.path.exists(file_path):
                return JsonResponse({'status': 'error', 'message': Messages.FILE_NOT_FOUND}, status=404)

            # metadata에서 columnLine 정보 가져오기
            column_line = 0
            if os.path.exists(metadata_path):
                try: 
                    with open(metadata_path, "r", encoding="utf-8") as f:
                        metadata = json.load(f)
                    column_line = metadata.get("columnLine", 0)
                except Exception as e:
                    print(f"Metadata 로드 오류: {e}")

            try:
                # openpyxl로 워크북 로드
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                
                # "집행내역" 시트 확인
                if WORKSHEET_NAME not in workbook.sheetnames:
                    return JsonResponse({
                        'status': 'error',
                        'data': None,
                        'message': Messages.WORKSHEET_NOT_FOUND.format(WORKSHEET_NAME),
                        'core_columns': core_columns,
                        'actual_columns': [],
                    }, status=400, safe=False, json_dumps_params={'ensure_ascii': False})
                
                worksheet = workbook[WORKSHEET_NAME]
                
                # 헤더 행 결정 (columnLine 기준)
                header_row = column_line + 1 if column_line >= 0 else 1
                
                # 헤더 읽기
                headers = []
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=header_row, column=col_idx).value
                    if cell_value is not None:
                        headers.append(str(cell_value).strip())
                    else:
                        headers.append(f"Column_{col_idx}")
                
                # 데이터 읽기 (헤더 다음 행부터)
                data_rows = []
                for row_idx in range(header_row + 1, worksheet.max_row + 1):
                    row_data = {}
                    has_data = False
                    
                    for col_idx, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None:
                            has_data = True
                            row_data[header] = cell_value
                        else:
                            row_data[header] = None
                    
                    # 빈 행이 아닌 경우만 추가
                    if has_data:
                        data_rows.append(row_data)
                
                # DataFrame과 유사한 구조로 변환
                actual_columns = headers
                
            except Exception as e:
                return JsonResponse({
                    'status': 'error',
                    'data': None,
                    'message': Messages.EXCEL_FILTER_ERROR,
                    'error_detail': str(e),
                    'core_columns': core_columns,
                    'actual_columns': [],
                }, status=500, safe=False, json_dumps_params={'ensure_ascii': False})

            # 핵심 컬럼이 없으면 생성
            for col in core_columns:
                if col not in actual_columns:
                    actual_columns.append(col)
                    # 모든 데이터 행에 해당 컬럼 추가
                    for row_data in data_rows:
                        row_data[col] = None

            # 핵심 컬럼 중 없는 것들 체크
            missing_core = [col for col in core_columns if col not in [h for h in headers if h in actual_columns]]
            if missing_core:
                return JsonResponse({
                    'status': 'warning',
                    'message': Messages.CORE_COLUMNS_MISSING.format(", ".join(missing_core)),
                    'data': data_rows,
                    'core_columns': core_columns,
                    'actual_columns': actual_columns,
                    'missing_core': missing_core
                }, status=200, safe=False, json_dumps_params={'ensure_ascii': False})

            return JsonResponse({
                'status': 'success',
                'data': data_rows,
                'message': Messages.EXCEL_READ_SUCCESS,
                'core_columns': core_columns,
                'actual_columns': actual_columns,
                'total_columns': len(actual_columns)
            }, status=200, safe=False, json_dumps_params={'ensure_ascii': False})

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'data': None,
                'message': str(e),
                'core_columns': core_columns,
                'actual_columns': [],
            }, status=500, safe=False, json_dumps_params={'ensure_ascii': False})

    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method',
        'data': None,
        'core_columns': core_columns,
        'actual_columns': [],
    }, status=405, safe=False, json_dumps_params={'ensure_ascii': False})

@csrf_exempt
def save_xlsx(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            folder_name = body.get("folderName")
            xlsx_file = body.get("xlsxFile")
            last_modified = body.get("lastModified")
            data = body.get("data")

            if not folder_name or not xlsx_file or not data:
                return JsonResponse({"status": "error", "message": Messages.MISSING_REQUIRED_INFO}, status=400)

            # 검토사항이 빈값이면 항상 빈 문자열로 저장
            for row in data:
                review = row.get(PROGRESS_COLUMN)
                if not review or (isinstance(review, list) and len(review) == 0):
                    row[PROGRESS_COLUMN] = ""
                elif isinstance(review, list):
                    row[PROGRESS_COLUMN] = ", ".join(str(x) for x in review)
                else:
                    row[PROGRESS_COLUMN] = str(review)

            file_path = os.path.join(UPLOAD_DIR, folder_name, xlsx_file)
            metadata_path = os.path.join(UPLOAD_DIR, folder_name, METADATA_FILENAME)

            # 메타데이터에서 columnLine 정보 가져오기
            column_line = 0
            if os.path.exists(metadata_path):
                with open(metadata_path, "r", encoding="utf-8") as f:
                    metadata = json.load(f)
                column_line = metadata.get("columnLine", 0)
            else:
                metadata = {}

            # DataFrame 생성
            df = pd.DataFrame(data)
            
            # columnLine 정보에 따라 저장 방식 결정
            if os.path.exists(file_path):
                # 기존 파일 구조 유지하면서 데이터만 업데이트
                workbook = openpyxl.load_workbook(file_path)
                
                # "집행내역" 시트 확인
                if WORKSHEET_NAME not in workbook.sheetnames:
                    return JsonResponse({"status": "error", "message": Messages.WORKSHEET_NOT_FOUND.format(WORKSHEET_NAME)}, status=400)
                
                worksheet = workbook[WORKSHEET_NAME]

                # 기존 데이터 영역 삭제 (헤더 아래부터)
                max_row = worksheet.max_row
                if max_row > column_line + 1:
                    worksheet.delete_rows(column_line + 2, max_row - column_line - 1)
                
                # 새 데이터 입력 (헤더 다음 행부터)
                for row_idx, (_, row_data) in enumerate(df.iterrows()):
                    for col_idx, value in enumerate(row_data):
                        worksheet.cell(
                            row=column_line + 2 + row_idx,
                            column=col_idx + 1,
                            value=value
                        )
                
                workbook.save(file_path)

            # openpyxl로 진행도 계산
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                if WORKSHEET_NAME in workbook.sheetnames:
                    worksheet = workbook[WORKSHEET_NAME]
                    metadata["progress"] = calculate_progress(worksheet, column_line)
                else:
                    metadata["progress"] = 0
            except Exception as e:
                print(f"진행도 계산 오류: {e}")
                metadata["progress"] = 0

            metadata["lastModified"] = last_modified or datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            with open(metadata_path, "w", encoding="utf-8") as f:
                json.dump(metadata, f, ensure_ascii=False, indent=4)

            return JsonResponse({"status": "success", "message": Messages.SAVE_SUCCESS})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

@csrf_exempt
def upload_rules(request):
    if request.method == "POST":
        upload_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rule_name = request.POST.get("rule_name")
        rule_dir = os.path.join(RULE_DIR, rule_name)
        os.makedirs(rule_dir, exist_ok=True)

        metadata = {
            "folderName": rule_name,
            "documentRule": None,
            "categoryRule": None,
            "uploadTime": upload_time,
        }

        fs = FileSystemStorage(location=rule_dir)

        # documentRule 저장 및 categoryRule 동적 생성
        if "document_rule" in request.FILES:
            rule_file = request.FILES["document_rule"]
            rule_path = fs.save(rule_file.name, rule_file)
            try:
                with open(os.path.join(rule_dir, rule_file.name), "r", encoding="utf-8") as f:
                    document_rule = json.load(f)
                metadata["documentRule"] = rule_file.name
                # document_rule의 key값을 기반으로 categoryRule 생성
                metadata["categoryRule"] = {key: list(document_rule[key].keys()) for key in document_rule}
            except Exception as e:
                metadata["documentRule"] = rule_file.name
                metadata["categoryRule"] = {}

        # metadata.json 저장
        with open(os.path.join(rule_dir, METADATA_FILENAME), "w", encoding="utf-8") as f:
            json.dump(metadata, f, ensure_ascii=False, indent=4)

        return JsonResponse({"message": Messages.RULE_UPLOAD_SUCCESS, "metadata": metadata})

    return JsonResponse({"error": Messages.INVALID_REQUEST_METHOD}, status=400)

def download_rule_zip(request, folder_name):
    folder_path = os.path.join(RULE_DIR, folder_name)
    if not os.path.isdir(folder_path):
        raise Http404("Rule folder not found")
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for root, _, files in os.walk(folder_path):
            for f in files:
                abs_path = os.path.join(root, f)
                rel_path = os.path.relpath(abs_path, folder_path)
                zip_file.write(abs_path, arcname=rel_path)
    zip_buffer.seek(0)
    zip_filename = f"{folder_name}.zip"
    return FileResponse(zip_buffer, as_attachment=True, filename=zip_filename)

@csrf_exempt
def save_rule(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            folder_name = body.get("folderName")
            document_rule = body.get("documentRule")

            if not folder_name:
                return JsonResponse({"status": "error", "message": Messages.MISSING_FOLDER_NAME}, status=400)

            rule_folder = os.path.join(RULE_DIR, folder_name)
            os.makedirs(rule_folder, exist_ok=True)

            # documentRule 저장
            if document_rule is not None:
                with open(os.path.join(rule_folder, document_rule), "w", encoding="utf-8") as f:
                    json.dump(document_rule, f, ensure_ascii=False, indent=4)

            # metadata.json의 lastModified만 업데이트
            metadata_path = os.path.join(rule_folder, METADATA_FILENAME)
            if os.path.exists(metadata_path):
                with open(metadata_path, "r", encoding="utf-8") as f:
                    metadata = json.load(f)
            else:
                metadata = {}
            metadata["lastModified"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(metadata_path, "w", encoding="utf-8") as f:
                json.dump(metadata, f, ensure_ascii=False, indent=4)

            return JsonResponse({"status": "success", "message": Messages.RULE_SAVE_SUCCESS})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

@csrf_exempt
def update_rule_name(request):
    """
    파일의 metadata.json에서 ruleName만 변경하는 함수
    """
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            folder_name = body.get("folderName")
            rule_name = body.get("ruleName")

            if not folder_name or rule_name is None:
                return JsonResponse({"status": "error", "message": Messages.MISSING_FOLDER_RULE_NAME}, status=400)

            metadata_path = os.path.join(UPLOAD_DIR, folder_name, METADATA_FILENAME)
            if not os.path.exists(metadata_path):
                return JsonResponse({"status": "error", "message": Messages.METADATA_NOT_EXISTS}, status=404)

            with open(metadata_path, "r", encoding="utf-8") as f:
                metadata = json.load(f)

            metadata["ruleName"] = rule_name

            with open(metadata_path, "w", encoding="utf-8") as f:
                json.dump(metadata, f, ensure_ascii=False, indent=4)

            return JsonResponse({"status": "success", "message": Messages.RULE_NAME_UPDATE_SUCCESS})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

from django.views.decorators.http import require_http_methods

@csrf_exempt
@require_http_methods(["DELETE"])
def delete_file(request, folder_name):
    """
    업로드된 파일 폴더(및 내부 파일 전체) 삭제
    """
    folder_path = os.path.join(UPLOAD_DIR, folder_name)
    if not os.path.exists(folder_path):
        return JsonResponse({"error": Messages.FOLDER_NOT_EXISTS}, status=404)
    try:
        import shutil
        shutil.rmtree(folder_path)
        return JsonResponse({"status": "success", "message": Messages.DELETE_SUCCESS.format(folder_name)})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
def delete_rule(request, folder_name):
    """
    업로드된 규칙 폴더(및 내부 파일 전체) 삭제
    """
    rule_folder_path = os.path.join(RULE_DIR, folder_name)
    if not os.path.exists(rule_folder_path):
        return JsonResponse({"error": Messages.RULE_FOLDER_NOT_EXISTS}, status=404)
    try:
        import shutil
        shutil.rmtree(rule_folder_path)
        return JsonResponse({"status": "success", "message": Messages.RULE_DELETE_SUCCESS.format(folder_name)})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)